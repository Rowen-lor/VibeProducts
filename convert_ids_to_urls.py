import openpyxl

def convert_ids_to_urls():
    """
    Reads IDs from the first column of 'Status_Inspection.xlsx',
    converts them to URLs, and saves them to column B of the same file.
    """
    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook('Status_Inspection.xlsx')
        sheet = workbook.active

        # Iterate over the rows in the first column of the source sheet
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
            if row[0]:
                original_id = row[0]
                # Construct the URL
                url = f"http://url.example.com/details?taskId={original_id}&pid=51&locale=en-US"
                # Write the URL to column B of the same row
                sheet.cell(row=row_index, column=2, value=url)

        # Save the workbook (overwrite original file)
        workbook.save('Status_Inspection.xlsx')
        print("Successfully updated 'Status_Inspection.xlsx' with URLs in column B.")

    except FileNotFoundError:
        print("Error: 'Status_Inspection.xlsx' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# The main execution block is removed to allow this script to be imported as a module.
# The convert_ids_to_urls() function will be called from the GUI application.
